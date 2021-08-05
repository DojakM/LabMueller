import re
from collections import deque
import glob
import openpyxl as px

class BabbleFish:
    def __init__(self):
        self.file = open("/Volumes/MN001122/Material Listen/translator.txt", 'r')

    def check_translation(self, word):
        for line in self.file:
            if re.search(".*" + word + ".*"):
                self.translations = line.split(",")
                return self.translations[0]
            else:
                return word

    def all_translations(self):
        return self.translations

class Construct:
    def __init__(self, promotor, sequence, fluor):
        self.promo = promotor
        self.seq = sequence
        self.rave = fluor
        self.seedline = []

    def add_seedline(self, seed_line):
        self.seedline.append(seed_line)

    def read_all_seedline(self):
        for i in self.seedline:
            print(i.name)

    def read_last_seedline(self):
        print(self.seedline[-1])


class Seed_line:
    def __init__(self, Name, Parents, Construct, date):
        self.name = Name
        self.parent = Parents
        self.construct = Construct
        self.date = date
    # getter methods
    def get_name(self):
        return self.name

    # setter method
    def set_name(self, newname):
        self.name = newname

    ## object methods
    # generate a new seedline
    def generate_seedline(self, amount):
        self.llist_of_seeds = deque()
        for i in range(amount):
            new_seed = Seed(self, i + 1)
            self.llist_of_seeds.append(new_seed)

    # read preexisting seedline, atm in console
    def read_seedline(self):
        try:
            for seed in self.llist_of_seeds:
                print(str(seed.get_ID()) + "\t|\t" + seed.get_status())
                print("---------------")
        except:
            print("NO RECORD")

    def generate_incoming_node(self):
        return ((self.parent, self.name))


class Seed(Seed_line):
    def __init__(self, seedline, ID, status="nothing"):
        self.parent = seedline
        self.ID = ID
        self.status = status

    # getter methods
    def get_ID(self):
        return self.ID

    def get_status(self):
        return self.status

    # setter methods
    def set_ID(self, newID):
        self.ID = newID

    def set_status(self, new_status):
        self.status = new_status

class ProteinSeq:
    def __init__(self, code, sequence = "", Class_group = "", Ath_orthologue = "",  motifs = ""):
        self.code = code
        self.seq = sequence
        self.Class = Class_group
        self.orth = Ath_orthologue
        self.motifs = motifs

    def save_me(self, changer):
        wb = px.load_workbook("/Volumes/MN001122/Material Listen/sequence.xlsx")
        ws = wb["Sequences"]
        for i in range(2, 100, 1):
            ind = "A" + str(i)
            if ws[ind].value == None or ws[ind].value == self.code:
                ws[ind] = self.code
                ws["B" + str(i)] = self.seq
                ws["C" + str(i)] = self.Class
                ws["D" + str(i)] = self.orth
                ws["E" + str(i)] = self.motifs
                ws["F" + str(i)] = changer
                break
        wb.save("/Volumes/MN001122/Material Listen/sequence.xlsx")

    def load_me(self):
        wb = px.load_workbook("/Volumes/MN001122/Material Listen/sequence.xlsx")
        ws = wb["Sequences"]
        for i in range(2, 100, 1):
            ind = "A" + str(i)
            ind2 = "G" + str(i)
            if (ws[ind].value == self.code) or (ws[ind2].value == self.code):
                self.seq = ws["B" + str(i)].value
                self.Class = ws["C" + str(i)].value
                self.orth = ws["D" + str(i)].value
                self.motifs = ws["E" + str(i)].value
                break

class Construct_reader():
    def __init__(self, protein_name):
        self.proteinName = protein_name
        self.constructs = []
        self.checkForConstructs()

    def checkForConstructs(self):
        URL = glob.glob("/Volumes/MN001122/Pantelis Livanos/Progress report/POKs overview table.xlsx")[0]
        wb = px.load_workbook(URL)
        ws = wb["Tabelle1"]
        i = 3
        while(True):
            coordinate = "A" + str(i)
            if ws[coordinate].value is None:
                break
            if re.match(".*" + self.proteinName + ".*", ws[coordinate].value):
                self.constructs.append(ws[coordinate].value)
            i = i+1