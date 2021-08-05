import glob
import pickle as pk
import re
import tkinter as tk
from tkinter import ttk
import openpyxl as px


# Basic picklers, URL has to be changed. IN future fix to a folder
def load_pkl(name):
    with open("./" + name + '.pkl', 'rb') as f:
        return pk.load(f)

def save_pkl(obj, name):
    with open("./" + name + ".pkl", "wb") as f:
        pk.dump(obj, f)

class Seedbox_list:
    #Constructor of this Window
    def __init__(self):

        #creation of a basic window
        cs_root = tk.Tk()
        cs_root.geometry("1024x420")

        #layout options
        cs_root.configure(bg="floral white")

        #treeview Constructor + layout
        tree = ttk.Treeview(cs_root, columns = (1,2,3,4,5), show="headings")
        tree.grid(row=4, column= 0, columnspan=4, sticky=tk.S + tk.W +tk.E)
        style = ttk.Style()
        style.theme_use("clam")
        style.configure("Treeview",
                        background="white",
                        foreground = "black",
                        rowheigth = 80,
                        fieldbackground = "white")

        #tree columns
        tree.heading("#1", text = "Construct")
        tree.heading("#2", text = "newest Seedline")
        tree.heading("#3", text = "Modification")
        tree.heading("#4", text = "Harvest Date")
        tree.heading("#5", text = "Box + position")

        #Search function setup
        lab = tk.StringVar()
        lab.set("Search:")
        tlabel = tk.Label(cs_root, textvariable=lab, bg="floral white")
        tlabel.grid(row=0, column=0)
        tfield = tk.Entry(cs_root)
        tfield.grid(row=1, column=0)

        #Search function
        def search_in_table():
            search_all_POKs()
            searchterm = tfield.get()
            selections = []
            for child in tree.get_children():
                if not re.search(".*" + searchterm + ".*", str(tree.item(child)['values'][0])):
                    tree.detach(child)

        def search_all_POKs():
                for child in tree.get_children():
                    if re.search(".*[pP]+[oO]+[kK]+[1]+.*", str(tree.item(child)['values'][0])):
                        file = open("POK1.txt", 'a')
                        file.write(str(tree.item(child)['values'][0]) + "\n")
                    if re.search(".*[pP]+[oO]+[kK]+[2]+.*", str(tree.item(child)['values'][0])):
                        file = open("POK2.txt", 'a')
                        file.write(str(tree.item(child)['values'][0]) + "\n")

        #loading from Excel
        def loading_from_Excel():
            URL = glob.glob("/Volumes/MN001122/Material Listen/seed_stock_safe*")
            wb = px.load_workbook(URL[0])
            ws = wb["seeds 2008"]
            a = 2
            while (True):
                harvest_date = ws["B" + str(a)].value
                seed_number = str(ws["C" + str(a)].value)
                gen_typ = ws["D" + str(a)].value
                box_index = ws["E" + str(a)].value
                position = ws["F" + str(a)].value
                modifications = ws["G" + str(a)].value
                if (gen_typ == "STOP"):
                    break
                elif seed_number == None:
                    a +=1
                elif re.search("[A-Z]+[1-9]+.*", seed_number):
                    tree.insert(parent='', index='end', values=(gen_typ, seed_number, modifications, harvest_date,
                                                                str(box_index)+ ": " + str(position)))
                    a += 1
                else:
                    a += 1

        search_button = tk.Button(cs_root, command=lambda: search_in_table(), text="GO!", bg="steel blue")
        search_button.grid(row=1, column=1, sticky="W")

        load_button = tk.Button(cs_root, command=lambda: loading_from_Excel(), text="Load from Excel", bg="steel blue")
        load_button.grid(row=1, column=2, sticky="W")


        tk.mainloop()

class Seedbox_box:
    def __init__(self):
        main = tk.Tk()
        main.title("Seedboxes")
        main.geometry("480x400")
        seed_boxes = load_pkl("boxes")
        row = 1
        col = 0
        for i in range(len(seed_boxes)+1):
            if col == 6:
                col = 0
                row += 1
            col += 1
            tk.Button(main, command=lambda: print("press"), text="Box {}".format(i)).grid(row = row, column = col)
        main.mainloop()

class run_through_excel():
    def __init__(self):
        URL = glob.glob("/Volumes/MN001122/Material Listen/seed_stock_safe*")
        wb = px.load_workbook(URL[0])
        ws = wb["seeds 2008"]
        a = 2
        box_bib = {}
        seed_boxes = []
        box = {}
        temp = 1
        while (True):
            gen_typ = ws["D" + str(a)].value
            box_index = ws["E" + str(a)].value
            position = ws["F" + str(a)].value
            if box_index == None:
                print(a)
            elif box_index == "box #":
                print(a)
            elif box_index != temp:
                seed_boxes.append(box)
                box = {}
                temp = box_index
            if (gen_typ == "STOP"):
                break
            elif gen_typ == "Genotype":
                a += 1
            elif gen_typ != None:
                box.update({position: gen_typ})
                box_bib.update({gen_typ: position})
                a += 1
            else:
                a += 1
        save_pkl(box_bib, "boxbib")
        save_pkl(seed_boxes, "boxes")

if __name__ == '__main__':
    e = Seedbox_list()